from datetime import datetime, timedelta
from flask import Flask, request, jsonify
import smtplib, threading, time, os, io
import psycopg2
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

app = Flask(__name__)

# ── DATABASE (Postgres / Neon) ──────────────────────────────────────────────
DATABASE_URL = os.environ.get("DATABASE_URL", "")

def get_db():
    """Open a fresh connection to the shared Postgres database."""
    if not DATABASE_URL:
        raise RuntimeError("DATABASE_URL environment variable is not set")
    return psycopg2.connect(DATABASE_URL, sslmode="require")

# ── EMAIL / SCHEDULE CONFIG ─────────────────────────────────────────────────
MONTHLY_REPORT_EMAIL = os.environ.get("MONTHLY_REPORT_EMAIL", "")  # e.g. shaanwagh835@gmail.com

def ist_now():
    """Current time in IST (UTC+5:30), without relying on system timezone data."""
    return datetime.utcnow() + timedelta(hours=5, minutes=30)

TEAM_MEMBERS = [
    {"id": "abhishek",  "name": "Abhishek",  "email": "abhishek.kapur1@maersk.com"},
    {"id": "akshay",    "name": "Akshay",    "email": "akshay.mathur@maersk.com"},
    {"id": "ashesh",    "name": "Ashesh",    "email": "ashesh.garg@maersk.com"},
    {"id": "avisek",    "name": "Avisek",    "email": "avisek.nath@maersk.com"},
    {"id": "dhiraj",    "name": "Dhiraj",    "email": "dhiraj.singh@maersk.com"},
    {"id": "kamakhya",  "name": "Kamakhya",  "email": "kamakhya.kinkar@maersk.com"},
    {"id": "manish",    "name": "Manish",    "email": "manish.sambhar@maersk.com"},
    {"id": "mohini",    "name": "Mohini",    "email": "mohini.agarwal@maersk.com"},
    {"id": "nibedita",  "name": "Nibedita",  "email": "nibedita.basak@maersk.com"},
    {"id": "shantanu",  "name": "Shantanu",  "email": "shantanu.wagh@maersk.com"},
    {"id": "suresh",    "name": "Suresh",    "email": "suresh.verma@maersk.com"},
    {"id": "robby",     "name": "Robby",     "email": "roby.jacob@maersk.com"},
    {"id": "milind",    "name": "Milind",    "email": "milind.sardar1@maersk.com"},
    {"id": "yashodip",  "name": "Yashodip",  "email": "yashodip.patil@maersk.com"},
    {"id": "sanju",     "name": "Sanju",     "email": "sanju.sasidharan@maersk.com"},
]
TOTAL_SEATS = 11

def init_db():
    conn = get_db()
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS bookings (
        id SERIAL PRIMARY KEY,
        name TEXT, date TEXT, seat INTEGER)""")
    # Safe migration: add the status column if it doesn't exist yet (won't touch existing data)
    c.execute("ALTER TABLE bookings ADD COLUMN IF NOT EXISTS status TEXT DEFAULT 'Office'")
    c.execute("UPDATE bookings SET status='Office' WHERE status IS NULL AND seat IS NOT NULL")
    c.execute("UPDATE bookings SET status='WFH' WHERE status IS NULL AND seat IS NULL")
    conn.commit()
    conn.close()

init_db()

def get_bookings_for_date(date_str):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT name, seat, status FROM bookings WHERE date=%s", (date_str,))
    rows = c.fetchall()
    conn.close()
    return rows

def send_email(subject, html_body, recipients, attachment_bytes=None, attachment_name=None):
    """Generic email sender used by both the daily reminder and the monthly roster."""
    EMAIL_USER = os.environ.get("EMAIL_USER", "")
    EMAIL_PASS = os.environ.get("EMAIL_PASS", "")
    if not EMAIL_USER or not EMAIL_PASS:
        print("EMAIL_USER or EMAIL_PASS not set.")
        return False
    if not recipients:
        print("No recipients to send to.")
        return False
    msg = MIMEMultipart("mixed")
    msg["Subject"] = subject
    msg["From"] = f"SeatSync <{EMAIL_USER}>"
    msg["To"] = ", ".join(recipients)
    alt = MIMEMultipart("alternative")
    alt.attach(MIMEText(html_body, "html"))
    msg.attach(alt)
    if attachment_bytes is not None:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment_bytes)
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f'attachment; filename="{attachment_name}"')
        msg.attach(part)
    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(EMAIL_USER, EMAIL_PASS)
            server.sendmail(EMAIL_USER, recipients, msg.as_string())
        print(f"Email '{subject}' sent at {datetime.now()}")
        return True
    except Exception as e:
        print(f"Email failed: {e}")
        return False

def send_daily_seat_reminder():
    """At 4:30 PM: checks tomorrow's seat availability and emails ONLY the people
    who haven't booked/selected any status yet for tomorrow."""
    tomorrow = ist_now().date() + timedelta(days=1)
    date_str = tomorrow.strftime("%Y-%m-%d")
    rows = get_bookings_for_date(date_str)
    office_count = sum(1 for name, seat, status in rows if status == "Office")
    already_responded = {name for name, seat, status in rows}
    available = TOTAL_SEATS - office_count
    recipients = [m["email"] for m in TEAM_MEMBERS if m["name"] not in already_responded]
    if not recipients:
        print(f"Everyone has already responded for {date_str}, no reminder needed.")
        return
    pretty_date = tomorrow.strftime("%A, %d %b %Y")
    if available > 0:
        subject = f"🟢 {available} seat(s) available for {pretty_date} — book soon!"
        html_body = f"""<h2>SeatSync — Seat Availability</h2>
        <p>You haven't selected your plan for <b>{pretty_date}</b> yet.</p>
        <p><b>{available}</b> out of {TOTAL_SEATS} seats are still available.</p>
        <p>Please book as soon as possible if you plan to come to office.</p>"""
    else:
        subject = f"🔴 No seats available for {pretty_date}"
        html_body = f"""<h2>SeatSync — Seat Availability</h2>
        <p>You haven't selected your plan for <b>{pretty_date}</b> yet.</p>
        <p>All {TOTAL_SEATS} seats are already booked.</p>
        <p>Sorry about that — please try to work from home tomorrow.</p>"""
    send_email(subject, html_body, recipients)

def build_monthly_roster_xlsx(year, month):
    """Builds an Excel roster: one row per weekday of the month, one column per team member,
    showing their status (Office/WFH/Travel/Leave/blank) for that day."""
    import calendar
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT name, date, seat, status FROM bookings WHERE date LIKE %s", (f"{year:04d}-{month:02d}-%",))
    rows = c.fetchall()
    conn.close()
    bookings_by_date = {}
    for name, date_str, seat, status in rows:
        bookings_by_date.setdefault(date_str, {})[name] = (status, seat)

    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Roster"
    header = ["Date", "Day"] + [m["name"] for m in TEAM_MEMBERS]
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="0073AB", end_color="0073AB", fill_type="solid")

    fills = {
        "Office": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "WFH":    PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "Travel": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
        "Leave":  PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid"),
        "":       PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid"),
    }

    days_in_month = calendar.monthrange(year, month)[1]
    for day in range(1, days_in_month + 1):
        d = datetime(year, month, day)
        if d.weekday() >= 5:  # skip weekends
            continue
        date_str = d.strftime("%Y-%m-%d")
        row_vals = [date_str, d.strftime("%A")]
        statuses_for_row = []
        for m in TEAM_MEMBERS:
            status, seat = bookings_by_date.get(date_str, {}).get(m["name"], ("", None))
            label = f"Office (Seat {seat})" if status == "Office" and seat else (status or "Not marked")
            row_vals.append(label)
            statuses_for_row.append(status)
        ws.append(row_vals)
        row_idx = ws.max_row
        for col_idx, status in enumerate(statuses_for_row, start=3):
            ws.cell(row=row_idx, column=col_idx).fill = fills.get(status, fills[""])

    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = max(12, max_len + 2)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def send_monthly_roster():
    today = ist_now().date()
    last_day = today  # this function is called ON the last day of the month
    recipients = [MONTHLY_REPORT_EMAIL] if MONTHLY_REPORT_EMAIL else []
    if not recipients:
        print("MONTHLY_REPORT_EMAIL not set, skipping monthly roster email.")
        return
    xlsx_bytes = build_monthly_roster_xlsx(last_day.year, last_day.month)
    month_name = last_day.strftime("%B %Y")
    subject = f"📊 SeatSync Monthly Roster — {month_name}"
    html_body = f"<h2>SeatSync Monthly Roster</h2><p>Attached is the full office/WFH roster for <b>{month_name}</b>.</p>"
    filename = f"SeatSync_Roster_{last_day.strftime('%Y_%m')}.xlsx"
    send_email(subject, html_body, recipients, xlsx_bytes, filename)

def email_scheduler():
    """Background loop: sends the daily 9PM IST seat reminder, and the monthly
    roster email on the last day of each month."""
    last_daily_sent = None
    last_monthly_sent = None
    while True:
        try:
            now = ist_now()
            today = now.strftime("%Y-%m-%d")
            tomorrow = (now + timedelta(days=1)).date()

            # Daily reminder at 16:30 (4:30 PM) IST — only if tomorrow is a working day (Mon-Fri)
            if now.hour == 16 and now.minute == 30 and last_daily_sent != today:
                if tomorrow.weekday() < 5:
                    send_daily_seat_reminder()
                last_daily_sent = today

            # Monthly roster at 21:30 IST on the last day of the month
            if now.hour == 21 and now.minute == 30 and last_monthly_sent != today:
                if tomorrow.day == 1:  # tomorrow rolls into a new month => today is the last day
                    send_monthly_roster()
                last_monthly_sent = today
        except Exception as e:
            print(f"Scheduler error: {e}")
        time.sleep(60)

threading.Thread(target=email_scheduler, daemon=True).start()

HTML_PAGE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>SeatSync — Maersk Machinery Mumbai</title>
<link href="https://fonts.googleapis.com/css2?family=Syne:wght@400;600;700;800&family=DM+Sans:wght@300;400;500&display=swap" rel="stylesheet">
<style>
:root {
  --teal:#0073AB; --sky:#42B4E6; --mint:#00B5B1;
  --gold:#E8A020; --coral:#E05A2B; --dark:#0A1628;
  --glass:rgba(255,255,255,0.07); --glass2:rgba(255,255,255,0.12);
  --border:rgba(255,255,255,0.1); --tl:rgba(255,255,255,0.6); --tm:rgba(255,255,255,0.85);
}
*{margin:0;padding:0;box-sizing:border-box;}
body{font-family:'DM Sans',sans-serif;background:var(--dark);color:#fff;min-height:100vh;overflow-x:hidden;}
body::before{content:'';position:fixed;inset:0;background:radial-gradient(ellipse 80% 60% at 10% 10%,rgba(0,115,171,0.25) 0%,transparent 60%),radial-gradient(ellipse 60% 50% at 90% 80%,rgba(0,181,177,0.15) 0%,transparent 55%);pointer-events:none;z-index:0;}
header{position:sticky;top:0;z-index:100;background:rgba(10,22,40,0.9);backdrop-filter:blur(20px);border-bottom:1px solid var(--border);padding:0 2rem;height:64px;display:flex;align-items:center;justify-content:space-between;}
.logo{display:flex;align-items:center;gap:12px;font-family:'Syne',sans-serif;font-weight:800;font-size:1.25rem;}
.logo-icon{width:36px;height:36px;background:linear-gradient(135deg,var(--teal),var(--mint));border-radius:10px;display:flex;align-items:center;justify-content:center;font-size:1.1rem;}
.logo span{color:var(--sky);}
.hbadge{background:var(--glass2);border:1px solid var(--border);border-radius:20px;padding:6px 14px;font-size:0.8rem;color:var(--tm);}
.hbadge strong{color:var(--sky);}
.app{position:relative;z-index:1;display:grid;grid-template-columns:280px 1fr;min-height:calc(100vh - 64px);}
.sidebar{background:rgba(10,22,40,0.6);border-right:1px solid var(--border);padding:1.5rem;display:flex;flex-direction:column;gap:1.5rem;overflow-y:auto;}
.stitle{font-family:'Syne',sans-serif;font-size:0.7rem;font-weight:600;letter-spacing:0.12em;text-transform:uppercase;color:var(--tl);margin-bottom:0.5rem;}
.wai{background:var(--glass);border:1px solid var(--border);border-radius:14px;padding:1rem;}
.usel{width:100%;background:rgba(255,255,255,0.05);border:1px solid rgba(255,255,255,0.15);border-radius:10px;color:#fff;font-family:'DM Sans',sans-serif;font-size:0.95rem;padding:10px 12px;outline:none;cursor:pointer;appearance:none;background-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='12' height='8' viewBox='0 0 12 8'%3E%3Cpath d='M1 1l5 5 5-5' stroke='%2342B4E6' stroke-width='1.5' fill='none' stroke-linecap='round'/%3E%3C/svg%3E");background-repeat:no-repeat;background-position:right 12px center;padding-right:32px;}
.usel option{background:#1a2a3a;color:#fff;}
.uav{width:44px;height:44px;border-radius:50%;display:flex;align-items:center;justify-content:center;font-family:'Syne',sans-serif;font-weight:700;font-size:1rem;margin-bottom:0.75rem;background:linear-gradient(135deg,#0073AB,#00B5B1);}
.mcal{background:var(--glass);border:1px solid var(--border);border-radius:14px;padding:1rem;}
.cnav{display:flex;align-items:center;justify-content:space-between;margin-bottom:0.75rem;}
.cmonth{font-family:'Syne',sans-serif;font-weight:700;font-size:0.9rem;}
.cbtn{background:var(--glass2);border:1px solid var(--border);border-radius:6px;color:var(--sky);width:26px;height:26px;display:flex;align-items:center;justify-content:center;cursor:pointer;font-size:0.9rem;transition:all 0.15s;}
.cbtn:hover{background:var(--teal);border-color:var(--teal);color:#fff;}
.cgrid{display:grid;grid-template-columns:repeat(7,1fr);gap:2px;}
.cdh{text-align:center;font-size:0.65rem;font-weight:500;color:var(--tl);padding:4px 0;}
.cd{aspect-ratio:1;display:flex;align-items:center;justify-content:center;font-size:0.75rem;border-radius:6px;cursor:pointer;position:relative;transition:all 0.15s;color:var(--tm);}
.cd:hover:not(.ce):not(.cp):not(.cw){background:var(--glass2);}
.ce,.cp{color:rgba(255,255,255,0.2);cursor:default;}
.cw{color:rgba(255,255,255,0.3);cursor:default;}
.chb::after{content:'';position:absolute;bottom:2px;width:4px;height:4px;border-radius:50%;background:var(--mint);}
.csel{background:var(--teal)!important;color:#fff!important;font-weight:600;}
.ctod{border:1px solid var(--sky);color:var(--sky);font-weight:600;}
.ctod.csel{border-color:transparent;background:var(--teal)!important;color:#fff!important;}
.legend{background:var(--glass);border:1px solid var(--border);border-radius:14px;padding:1rem;display:flex;flex-direction:column;gap:8px;}
.li{display:flex;align-items:center;gap:8px;font-size:0.8rem;color:var(--tm);}
.ld{width:10px;height:10px;border-radius:3px;flex-shrink:0;}
.stats{display:flex;flex-direction:column;gap:8px;}
.sc{background:var(--glass);border:1px solid var(--border);border-radius:12px;padding:0.75rem 1rem;display:flex;align-items:center;gap:10px;}
.si{font-size:1.2rem;width:32px;text-align:center;}
.sv{font-family:'Syne',sans-serif;font-weight:700;font-size:1.1rem;color:var(--sky);}
.sl{font-size:0.72rem;color:var(--tl);}
.rc{background:linear-gradient(135deg,rgba(232,160,32,0.1),rgba(232,160,32,0.05));border:1px solid rgba(232,160,32,0.25);border-radius:16px;padding:1.2rem 1.5rem;display:flex;align-items:center;gap:1rem;}
.rt{font-family:'Syne',sans-serif;font-weight:700;font-size:0.9rem;color:var(--gold);}
.rd{font-size:0.78rem;color:var(--tl);margin-top:2px;}
.main{padding:2rem;display:flex;flex-direction:column;gap:1.5rem;overflow-y:auto;}
.dh{display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:1rem;}
.dtitle{font-family:'Syne',sans-serif;font-weight:800;font-size:1.8rem;letter-spacing:-0.03em;}
.dtitle span{color:var(--sky);}
.dsub{color:var(--tl);font-size:0.85rem;margin-top:4px;}
.abtns{display:flex;gap:0.75rem;}
.btn{font-family:'DM Sans',sans-serif;font-size:0.85rem;font-weight:500;border-radius:10px;padding:10px 18px;cursor:pointer;transition:all 0.2s;border:none;display:flex;align-items:center;gap:6px;}
.bp{background:linear-gradient(135deg,var(--teal),var(--mint));color:#fff;box-shadow:0 4px 15px rgba(0,181,177,0.3);}
.bp:hover{transform:translateY(-2px);box-shadow:0 6px 20px rgba(0,181,177,0.4);}
.bs{background:var(--glass2);border:1px solid var(--border);color:var(--tm);}
.bs:hover{background:rgba(255,255,255,0.15);color:#fff;}
.btn:disabled{opacity:0.4;cursor:not-allowed;transform:none!important;}
.ws{display:grid;grid-template-columns:repeat(5,1fr);gap:0.75rem;}
.wdc{background:var(--glass);border:1px solid var(--border);border-radius:14px;padding:1rem;cursor:pointer;transition:all 0.2s;position:relative;overflow:hidden;}
.wdc::before{content:'';position:absolute;top:0;left:0;right:0;height:3px;background:linear-gradient(90deg,var(--teal),var(--mint));opacity:0;transition:opacity 0.2s;}
.wdc:hover{border-color:rgba(0,115,171,0.4);transform:translateY(-2px);}
.wdc:hover::before{opacity:1;}
.wdc.act{background:rgba(0,115,171,0.2);border-color:var(--teal);}
.wdc.act::before{opacity:1;}
.wdcday{font-size:0.7rem;color:var(--tl);text-transform:uppercase;letter-spacing:0.08em;}
.wdcdate{font-family:'Syne',sans-serif;font-weight:700;font-size:1.4rem;margin:2px 0 6px;}
.wdcs{display:flex;align-items:center;gap:6px;font-size:0.78rem;}
.sbar{flex:1;height:4px;background:rgba(255,255,255,0.1);border-radius:2px;overflow:hidden;}
.sbf{height:100%;border-radius:2px;transition:width 0.4s ease;background:linear-gradient(90deg,var(--mint),var(--sky));}
.sbf.full{background:linear-gradient(90deg,var(--coral),#ff8c66);}
.bnames{margin-top:6px;display:flex;flex-wrap:wrap;gap:3px;}
.nc{font-size:0.62rem;padding:2px 6px;border-radius:4px;font-weight:500;border:1px solid rgba(255,255,255,0.1);}
.fps{background:var(--glass);border:1px solid var(--border);border-radius:20px;padding:1.5rem;}
.fph{display:flex;align-items:center;justify-content:space-between;margin-bottom:1.5rem;}
.fpt{font-family:'Syne',sans-serif;font-weight:700;font-size:1rem;}
.fpm{font-size:0.8rem;color:var(--tl);margin-top:2px;}
.fp{display:flex;flex-direction:column;gap:1.5rem;align-items:center;}
.fpl{font-size:0.7rem;text-transform:uppercase;letter-spacing:0.1em;color:var(--tl);text-align:center;margin-bottom:0.5rem;}
.sr{display:flex;gap:1rem;justify-content:center;flex-wrap:wrap;}
.seat{width:82px;aspect-ratio:1;border-radius:14px;display:flex;flex-direction:column;align-items:center;justify-content:center;cursor:pointer;transition:all 0.2s;border:2px solid transparent;user-select:none;}
.seat.free{background:rgba(0,181,177,0.1);border-color:rgba(0,181,177,0.35);}
.seat.free:hover{background:rgba(0,181,177,0.2);border-color:var(--mint);transform:translateY(-3px) scale(1.03);box-shadow:0 8px 20px rgba(0,181,177,0.25);}
.seat.taken{background:rgba(224,90,43,0.1);border-color:rgba(224,90,43,0.25);cursor:not-allowed;}
.seat.mine{background:rgba(232,160,32,0.15);border-color:rgba(232,160,32,0.5);}
.seat.mine:hover{background:rgba(232,160,32,0.2);border-color:var(--gold);transform:translateY(-3px);box-shadow:0 8px 20px rgba(232,160,32,0.2);}
.sicon{font-size:1.5rem;margin-bottom:2px;}
.snum{font-family:'Syne',sans-serif;font-weight:700;font-size:0.8rem;}
.seat.free .snum{color:var(--mint);}
.seat.taken .snum{color:rgba(255,255,255,0.4);}
.seat.mine .snum{color:var(--gold);}
.sname{font-size:0.6rem;color:rgba(255,255,255,0.75);text-align:center;max-width:74px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;padding:0 4px;}
.aisle{width:90%;height:1px;background:linear-gradient(90deg,transparent,var(--border),transparent);margin:0.5rem 0;display:flex;align-items:center;justify-content:center;}
.ailabel{background:var(--dark);padding:2px 10px;font-size:0.65rem;letter-spacing:0.1em;color:var(--tl);text-transform:uppercase;border:1px solid var(--border);border-radius:10px;}
.bt{background:var(--glass);border:1px solid var(--border);border-radius:20px;overflow:hidden;}
.bth{display:grid;grid-template-columns:140px repeat(5,1fr);background:rgba(0,115,171,0.15);border-bottom:1px solid var(--border);padding:0.75rem 1rem;font-size:0.72rem;font-weight:600;letter-spacing:0.06em;text-transform:uppercase;color:var(--tl);}
.btr{display:grid;grid-template-columns:140px repeat(5,1fr);padding:0.6rem 1rem;border-bottom:1px solid rgba(255,255,255,0.04);align-items:center;font-size:0.82rem;}
.btr:last-child{border-bottom:none;}
.btr:hover{background:var(--glass2);}
.btp{display:flex;align-items:center;gap:8px;font-weight:500;}
.btav{width:24px;height:24px;border-radius:50%;display:flex;align-items:center;justify-content:center;font-size:0.65rem;font-weight:700;flex-shrink:0;}
.btc{display:flex;align-items:center;justify-content:center;}
.pill{padding:3px 10px;border-radius:6px;font-size:0.72rem;font-weight:500;}
.pin{background:rgba(0,181,177,0.2);color:var(--mint);border:1px solid rgba(0,181,177,0.3);}
.pwfh{background:rgba(255,255,255,0.05);color:rgba(255,255,255,0.3);border:1px solid rgba(255,255,255,0.08);}
.ptravel{background:rgba(66,180,230,0.15);color:var(--sky);border:1px solid rgba(66,180,230,0.3);}
.pleave{background:rgba(224,90,43,0.15);color:var(--coral);border:1px solid rgba(224,90,43,0.3);}
.pnone{background:rgba(255,255,255,0.03);color:rgba(255,255,255,0.2);border:1px dashed rgba(255,255,255,0.12);}
.tabs{display:flex;gap:0.5rem;background:var(--glass);border:1px solid var(--border);border-radius:12px;padding:4px;width:fit-content;}
.tab{padding:7px 16px;border-radius:8px;cursor:pointer;font-size:0.82rem;font-weight:500;transition:all 0.15s;color:var(--tl);border:none;background:transparent;font-family:'DM Sans',sans-serif;}
.tab.act{background:var(--teal);color:#fff;}
.tab:hover:not(.act){color:#fff;}
.vp{display:none;}
.vp.act{display:block;animation:fi 0.25s ease;}
@keyframes fi{from{opacity:0;transform:translateY(8px);}to{opacity:1;transform:translateY(0);}}
.tc{position:fixed;bottom:1.5rem;right:1.5rem;z-index:1000;display:flex;flex-direction:column;gap:0.5rem;}
.toast{background:#1a2a3a;border:1px solid var(--border);border-radius:12px;padding:12px 16px;display:flex;align-items:center;gap:10px;font-size:0.85rem;box-shadow:0 8px 30px rgba(0,0,0,0.4);animation:ti 0.3s ease;max-width:320px;}
.toast.s{border-left:3px solid var(--mint);}
.toast.w{border-left:3px solid var(--gold);}
.toast.e{border-left:3px solid var(--coral);}
.toast.i{border-left:3px solid var(--sky);}
@keyframes ti{from{transform:translateX(100px);opacity:0;}to{transform:translateX(0);opacity:1;}}
.mo{position:fixed;inset:0;background:rgba(0,0,0,0.7);backdrop-filter:blur(8px);z-index:200;display:flex;align-items:center;justify-content:center;opacity:0;pointer-events:none;transition:opacity 0.2s;}
.mo.act{opacity:1;pointer-events:all;}
.modal{background:linear-gradient(135deg,#0d1f33,#132438);border:1px solid rgba(0,115,171,0.4);border-radius:20px;padding:2rem;max-width:440px;width:90%;transform:scale(0.95) translateY(10px);transition:transform 0.2s;box-shadow:0 20px 60px rgba(0,0,0,0.5);}
.mo.act .modal{transform:scale(1) translateY(0);}
.mtitle{font-family:'Syne',sans-serif;font-weight:700;font-size:1.2rem;margin-bottom:0.5rem;}
.mbody{color:var(--tm);font-size:0.9rem;line-height:1.6;margin-bottom:1.5rem;}
.mact{display:flex;gap:0.75rem;justify-content:flex-end;}
::-webkit-scrollbar{width:6px;}::-webkit-scrollbar-track{background:transparent;}::-webkit-scrollbar-thumb{background:rgba(255,255,255,0.1);border-radius:3px;}

/* ── MOBILE RESPONSIVE (phones & small tablets) ────────────────────────── */
@media (max-width: 860px) {
  header{padding:0 1rem;height:auto;min-height:56px;flex-wrap:wrap;gap:6px;padding-top:8px;padding-bottom:8px;}
  .logo{font-size:1.05rem;}
  .hbadge{font-size:0.7rem;padding:5px 10px;}
  .app{grid-template-columns:1fr;}
  .sidebar{border-right:none;border-bottom:1px solid var(--border);padding:1rem;}
  .main{padding:1rem;gap:1rem;}
  .dtitle{font-size:1.3rem;}
  .dh{flex-direction:column;align-items:flex-start;}
  .abtns{width:100%;}
  .abtns .btn{flex:1;justify-content:center;font-size:0.78rem;padding:10px 8px;}
  .tabs{width:100%;overflow-x:auto;}
  .tab{white-space:nowrap;padding:7px 12px;font-size:0.78rem;}
  .ws{grid-template-columns:repeat(2,1fr);gap:0.5rem;}
  .sr{gap:0.5rem;}
  .seat{width:64px;}
  .sicon{font-size:1.1rem;}
  .snum{font-size:0.68rem;}
  .sname{font-size:0.52rem;max-width:56px;}
  .bt{overflow-x:auto;}
  .bth,.btr{grid-template-columns:110px repeat(5,minmax(84px,1fr));min-width:530px;}
  .btp{font-size:0.78rem;}
  .modal{padding:1.25rem;width:92%;}
}
@media (max-width: 480px) {
  .logo-icon{width:30px;height:30px;font-size:0.95rem;}
  .dtitle{font-size:1.1rem;}
  .stats{gap:6px;}
  .sc{padding:0.6rem 0.75rem;}
  .seat{width:56px;}
  .fps{padding:1rem;}
}
</style>
</head>
<body>
<header>
  <div class="logo"><div class="logo-icon">🪑</div>Seat<span>Sync</span></div>
  <div style="display:flex;gap:1rem;">
    <div class="hbadge" id="wkbadge">Week of <strong>—</strong></div>
    <div class="hbadge">🏢 Maersk Machinery Mumbai · 11 Seats</div>
  </div>
</header>

<div class="app">
  <aside class="sidebar">
    <div>
      <div class="stitle">👤 Booking as</div>
      <div class="wai">
        <div class="uav" id="uav">?</div>
        <select class="usel" id="usel">
          <option value="">— Select your name —</option>
          <option value="abhishek">Abhishek</option>
          <option value="akshay">Akshay</option>
          <option value="ashesh">Ashesh</option>
          <option value="avisek">Avisek</option>
          <option value="dhiraj">Dhiraj</option>
          <option value="kamakhya">Kamakhya</option>
          <option value="manish">Manish</option>
          <option value="mohini">Mohini</option>
          <option value="nibedita">Nibedita</option>
          <option value="shantanu">Shantanu</option>
          <option value="suresh">Suresh</option>
          <option value="robby">Robby</option>
          <option value="milind">Milind</option>
          <option value="yashodip">Yashodip</option>
          <option value="sanju">Sanju</option>
        </select>
      </div>
    </div>

    <div>
      <div class="stitle">📅 Calendar</div>
      <div class="mcal">
        <div class="cnav">
          <button class="cbtn" id="prevM">‹</button>
          <div class="cmonth" id="cml">—</div>
          <button class="cbtn" id="nextM">›</button>
        </div>
        <div class="cgrid" id="cgrid"></div>
      </div>
    </div>

    <div>
      <div class="stitle">📊 Stats</div>
      <div class="stats">
        <div class="sc"><div class="si">🟢</div><div><div class="sv" id="stF">—</div><div class="sl">Available seats</div></div></div>
        <div class="sc"><div class="si">👥</div><div><div class="sv" id="stB">—</div><div class="sl">Booked today</div></div></div>
        <div class="sc"><div class="si">🏠</div><div><div class="sv" id="stW">—</div><div class="sl">Working from home</div></div></div>
      </div>
    </div>

    <div>
      <div class="stitle">Legend</div>
      <div class="legend">
        <div class="li"><div class="ld" style="background:var(--mint)"></div>Available seat</div>
        <div class="li"><div class="ld" style="background:var(--coral)"></div>Booked by someone</div>
        <div class="li"><div class="ld" style="background:var(--gold)"></div>Your booking</div>
      </div>
    </div>

    <div class="rc">
      <div style="font-size:1.8rem">⏰</div>
      <div><div class="rt">Daily Reminder</div><div class="rd" id="rdesc">Every day at 4:30 PM IST</div></div>
    </div>
  </aside>

  <main class="main">
    <div class="dh">
      <div>
        <div class="dtitle" id="dtitle">Select a <span>date</span></div>
        <div class="dsub" id="dsub">Click any working day in the calendar to get started</div>
      </div>
      <div class="abtns">
        <button class="btn bs" onclick="openWeekModal()">📅 Book Whole Week</button>
        <button class="btn bp" id="qbtn" onclick="quickBook()" disabled>✚ Book This Day</button>
      </div>
    </div>

    <div style="display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:1rem;">
      <div class="tabs">
        <button class="tab act" onclick="switchTab('floor',this)">🗺️ Floor Plan</button>
        <button class="tab" onclick="switchTab('week',this)">📋 Week View</button>
        <button class="tab" onclick="switchTab('roster',this)">👥 Full Roster</button>
      </div>
      <div style="font-size:0.78rem;color:var(--tl);" id="dbadge"></div>
    </div>

    <div class="ws" id="wstrip"></div>

    <div class="vp act" id="pf">
      <div class="fps">
        <div class="fph"><div><div class="fpt">Office Floor Plan</div><div class="fpm" id="fmeta">Select a date to view bookings</div></div></div>
        <div class="fp" id="fp"><div style="color:var(--tl);text-align:center;padding:3rem;font-size:0.9rem;">👈 Click any date in the calendar on the left to view seats</div></div>
      </div>
    </div>
    <div class="vp" id="pw"><div class="bt" id="wtbl"></div></div>
    <div class="vp" id="pr"><div class="bt" id="rtbl"></div></div>
  </main>
</div>

<div class="tc" id="tc"></div>
<div class="mo" id="mo">
  <div class="modal">
    <div class="mtitle" id="mtitle">Confirm</div>
    <div class="mbody" id="mbody"></div>
    <div class="mact" id="mact"></div>
  </div>
</div>

<script>
const TEAM=[
  {id:'abhishek',name:'Abhishek',color:'#4A90D9',ini:'AK'},
  {id:'akshay',  name:'Akshay',  color:'#7B68EE',ini:'AM'},
  {id:'ashesh',  name:'Ashesh',  color:'#20B2AA',ini:'AG'},
  {id:'avisek',  name:'Avisek',  color:'#FF6B6B',ini:'AN'},
  {id:'dhiraj',  name:'Dhiraj',  color:'#FFD700',ini:'DS'},
  {id:'kamakhya',name:'Kamakhya',color:'#FF8C00',ini:'KK'},
  {id:'manish',  name:'Manish',  color:'#32CD32',ini:'MS'},
  {id:'mohini',  name:'Mohini',  color:'#FF69B4',ini:'MA'},
  {id:'nibedita',name:'Nibedita',color:'#40E0D0',ini:'NB'},
  {id:'shantanu',name:'Shantanu',color:'#9370DB',ini:'SW'},
  {id:'suresh',  name:'Suresh',  color:'#E8544A',ini:'SV'},
  {id:'robby',   name:'Robby',   color:'#5CC8FF',ini:'RJ'},
  {id:'milind',  name:'Milind',  color:'#A3D977',ini:'MS'},
  {id:'yashodip',name:'Yashodip',color:'#C77DFF',ini:'YP'},
  {id:'sanju',   name:'Sanju',   color:'#FFB84D',ini:'SS'},
];
const SEATS=11, LAYOUT=[[1,2,3,4,5,6],[7,8,9,10,11]];

let S={
  user: localStorage.getItem('ss_user')||'',
  date: null,
  yr: new Date().getFullYear(),
  mo: new Date().getMonth(),
  cache: {},
  tab: 'floor'
};

// ── API ──────────────────────────────────────────────────────────────────
async function apiFetch(dateStr){
  if(S.cache[dateStr]!==undefined) return S.cache[dateStr];
  try{
    const r=await fetch('/api/bookings/'+dateStr);
    const d=await r.json();
    S.cache[dateStr]=d.bookings||[];
  }catch(e){S.cache[dateStr]=[];}
  return S.cache[dateStr];
}
async function apiBook(name,date,seat,status='Office'){
  const r=await fetch('/api/book',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({name,date,seat,status})});
  return r.json();
}
async function apiCancel(name,date){
  const r=await fetch('/api/cancel',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({name,date})});
  return r.json();
}
async function apiDelete(name,date,seat){
  const r=await fetch('/api/delete',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({name,date,seat})});
  return r.json();
}

// ── UTILS ────────────────────────────────────────────────────────────────
function dk(d){
  // Returns YYYY-MM-DD using LOCAL date (not UTC) to avoid timezone shifts
  const y=d.getFullYear(), m=String(d.getMonth()+1).padStart(2,'0'), day=String(d.getDate()).padStart(2,'0');
  return `${y}-${m}-${day}`;
}
function isWE(d){return d.getDay()===0||d.getDay()===6;}
function isPast(d){const t=new Date();t.setHours(0,0,0,0);const x=new Date(d);x.setHours(0,0,0,0);return x<t;}
function weekOf(d){
  const x=new Date(d),day=x.getDay(),diff=day===0?-6:1-day,mon=new Date(x);
  mon.setDate(x.getDate()+diff);
  return [0,1,2,3,4].map(i=>{const dd=new Date(mon);dd.setDate(mon.getDate()+i);return dd;});
}
function fmt(d,o){return d.toLocaleDateString('en-IN',o);}
function gm(id){return TEAM.find(t=>t.id===id);}
function getSeat(bkgs,uid){return bkgs.find(b=>b.name===uid)?.seat;}
function getEntry(bkgs,uid){return bkgs.find(b=>b.name===uid);}
function officeOf(bkgs){return bkgs.filter(b=>b.status==='Office');}
function statusPill(entry){
  if(!entry) return `<div class="pill pnone">—</div>`;
  if(entry.status==='Office') return `<div class="pill pin">S${entry.seat} 🏢</div>`;
  if(entry.status==='Travel') return `<div class="pill ptravel">✈️ Travel</div>`;
  if(entry.status==='Leave') return `<div class="pill pleave">🏖️ Leave</div>`;
  return `<div class="pill pwfh">🏠 WFH</div>`;
}
function bust(dateStr){delete S.cache[dateStr];}

// ── CALENDAR ─────────────────────────────────────────────────────────────
function renderCal(){
  const M=['January','February','March','April','May','June','July','August','September','October','November','December'];
  document.getElementById('cml').textContent=M[S.mo]+' '+S.yr;
  const today=new Date();today.setHours(0,0,0,0);
  const first=new Date(S.yr,S.mo,1);
  const dim=new Date(S.yr,S.mo+1,0).getDate();
  let h=['S','M','T','W','T','F','S'].map(x=>`<div class="cdh">${x}</div>`).join('');
  for(let i=0;i<first.getDay();i++) h+=`<div class="cd ce"></div>`;
  for(let d=1;d<=dim;d++){
    const date=new Date(S.yr,S.mo,d);
    const dkey=dk(date);
    const isT=date.getTime()===today.getTime();
    const isSel=S.date&&dk(S.date)===dkey;
    const past=isPast(date)&&!isT;
    const we=isWE(date);
    const hasB=(S.cache[dkey]||[]).length>0;
    let cls='cd';
    if(past) cls+=' cp'; else if(we) cls+=' cw';
    if(isT) cls+=' ctod';
    if(isSel) cls+=' csel';
    if(hasB&&!we) cls+=' chb';
    const click=!past&&!we;
    h+=`<div class="${cls}" ${click?`onclick="selDate(new Date(${S.yr},${S.mo},${d}))"`:''} >${d}</div>`;
  }
  document.getElementById('cgrid').innerHTML=h;
  if(S.date){
    const wd=weekOf(S.date);
    document.getElementById('wkbadge').innerHTML=`Week of <strong>${fmt(wd[0],{month:'short',day:'numeric'})}</strong>`;
  }
}

// ── SELECT DATE ───────────────────────────────────────────────────────────
async function selDate(date){
  S.date=date;
  await apiFetch(dk(date));
  renderCal();
  updateDH();
  await renderWS();
  await renderPanel();
  await updateStats();
}

function updateDH(){
  const d=S.date; if(!d) return;
  const dkey=dk(d);
  document.getElementById('dtitle').innerHTML=fmt(d,{weekday:'long'})+', <span>'+fmt(d,{day:'numeric',month:'short'})+'</span>';
  document.getElementById('dsub').textContent=fmt(d,{day:'numeric',month:'long',year:'numeric'});
  const b=S.cache[dkey]||[];
  document.getElementById('dbadge').textContent=(SEATS-officeOf(b).length)+' of '+SEATS+' seats free';
  document.getElementById('qbtn').disabled=!S.user||isWE(d)||isPast(d);
}

// ── WEEK STRIP ────────────────────────────────────────────────────────────
async function renderWS(){
  if(!S.date) return;
  const wd=weekOf(S.date);
  const dns=['Mon','Tue','Wed','Thu','Fri'];
  await Promise.all(wd.map(d=>apiFetch(dk(d))));
  document.getElementById('wstrip').innerHTML=wd.map((date,i)=>{
    const dkey=dk(date);
    const bkgs=S.cache[dkey]||[];
    const off=officeOf(bkgs);
    const free=SEATS-off.length;
    const pct=(off.length/SEATS)*100;
    const isSel=S.date&&dk(S.date)===dkey;
    const chips=bkgs.slice(0,4).map(b=>{const m=gm(b.name);return m?`<div class="nc" style="background:${m.color}22;color:${m.color};border-color:${m.color}33">${m.name.slice(0,3)}</div>`:''}).join('')+(bkgs.length>4?`<div class="nc" style="opacity:0.5">+${bkgs.length-4}</div>`:'');
    return `<div class="wdc ${isSel?'act':''}" onclick="selDate(new Date(${S.yr},${date.getMonth()},${date.getDate()}))">
      <div class="wdcday">${dns[i]}</div>
      <div class="wdcdate">${date.getDate()}</div>
      <div class="wdcs"><div class="sbar"><div class="sbf ${free===0?'full':''}" style="width:${pct}%"></div></div>
      <span style="font-size:0.7rem;color:${free===0?'var(--coral)':'var(--mint)'};white-space:nowrap">${free} free</span></div>
      ${chips?`<div class="bnames">${chips}</div>`:''}
    </div>`;
  }).join('');
}

// ── FLOOR PLAN ────────────────────────────────────────────────────────────
async function renderFloor(){
  const d=S.date;
  if(!d){
    document.getElementById('fp').innerHTML='<div style="color:var(--tl);text-align:center;padding:3rem;font-size:0.9rem;">👈 Click any date in the calendar to view seats</div>';
    return;
  }
  const dkey=dk(d);
  const bkgs=await apiFetch(dkey);
  const sm={};
  bkgs.filter(b=>b.status==='Office').forEach(b=>sm[b.seat]=b.name);
  document.getElementById('fmeta').textContent=fmt(d,{weekday:'long',day:'numeric',month:'long'})+' · '+officeOf(bkgs).length+'/'+SEATS+' booked';
  let h='';
  LAYOUT.forEach((row,ri)=>{
    h+=`<div><div class="fpl">${ri===0?'Row A':'Row B'}</div><div class="sr">`;
    row.forEach(sn=>{
      const uid=sm[sn], m=uid?gm(uid):null, isMe=uid===S.user;
      const cls=uid?(isMe?'mine':'taken'):'free';
      const canBook=!uid&&S.user&&!isPast(d)&&!isWE(d);
      const canCancel=isMe&&!isPast(d);
      const canDelete=uid&&!isPast(d);
      const oc=canBook?`onclick="bookSeat(${sn})"`:canCancel?`onclick="cancelSeat(${sn})"`:canDelete?`onclick="deleteBooking('${uid}', '${dkey}', ${sn})"`:'';      h+=`<div class="seat ${cls}" ${oc} title="${m?m.name:(canBook?'Click to book':'')}">
        <div class="sicon">${uid?(isMe?'⭐':'🧑'):'🪑'}</div>
        <div class="snum">S${sn}</div>
        <div class="sname">${m?m.name:'Free'}</div>
      </div>`;
    });
    h+=`</div></div>`;
    if(ri===0) h+=`<div class="aisle" style="width:100%"><div class="ailabel">aisle</div></div>`;
  });
  document.getElementById('fp').innerHTML=h;
}

// ── WEEK TABLE ────────────────────────────────────────────────────────────
async function renderWeek(){
  const d=S.date||new Date();
  const wd=weekOf(d);
  const dns=['Mon','Tue','Wed','Thu','Fri'];
  await Promise.all(wd.map(x=>apiFetch(dk(x))));
  let h=`<div class="bth"><div>Team Member</div>${wd.map((x,i)=>`<div style="text-align:center">${dns[i]}<br><span style="font-weight:400;color:var(--sky)">${fmt(x,{day:'numeric',month:'short'})}</span></div>`).join('')}</div>`;
  TEAM.forEach(m=>{
    h+=`<div class="btr"><div class="btp"><div class="btav" style="background:${m.color}33;color:${m.color}">${m.ini}</div>${m.name}</div>`;
    wd.forEach(x=>{
      const dkey = dk(x);
      const bkgs = S.cache[dkey]||[];
      const entry = getEntry(bkgs, m.id);
      const past = isPast(x);
      const isMe = m.id===S.user;
      if(isMe && !past){
        // Editable dropdown for the logged-in user's own row
        h+=`<div class="btc"><select class="usel" style="padding:4px 8px;font-size:0.75rem;width:auto;min-width:100px;" onchange="setStatus('${dkey}', this.value)">
          <option value="" ${!entry?'selected':''}>— Select —</option>
          <option value="Office" ${entry&&entry.status==='Office'?'selected':''}>🏢 Office</option>
          <option value="WFH" ${entry&&entry.status==='WFH'?'selected':''}>🏠 WFH</option>
          <option value="Travel" ${entry&&entry.status==='Travel'?'selected':''}>✈️ Travel</option>
          <option value="Leave" ${entry&&entry.status==='Leave'?'selected':''}>🏖️ Leave</option>
        </select></div>`;
      } else {
        const canDelete = entry && !past;
        h+=`<div class="btc" ${canDelete?`onclick="deleteBooking('${m.id}', '${dkey}')" style="cursor:pointer" title="Click to delete (admin)"`:''}>${statusPill(entry)}</div>`;
      }
    });
    h+=`</div>`;
  });
  document.getElementById('wtbl').innerHTML=h;
}

// ── ROSTER ────────────────────────────────────────────────────────────────
async function renderRoster(){
  const days=[];
  let x=new Date();
  while(days.length<5){if(!isWE(x)) days.push(new Date(x));x.setDate(x.getDate()+1);}
  await Promise.all(days.map(d=>apiFetch(dk(d))));
  let h=`<div class="bth" style="grid-template-columns:140px repeat(5,1fr)"><div>Member</div>${days.map(d=>`<div style="text-align:center">${fmt(d,{weekday:'short'})}<br><span style="font-weight:400;color:var(--sky)">${fmt(d,{day:'numeric',month:'short'})}</span></div>`).join('')}</div>`;
  TEAM.forEach(m=>{
    h+=`<div class="btr" style="grid-template-columns:140px repeat(5,1fr)"><div class="btp"><div class="btav" style="background:${m.color}33;color:${m.color}">${m.ini}</div>${m.name}</div>`;
    days.forEach(d=>{
      const dkey = dk(d);
      const entry=getEntry(S.cache[dkey]||[],m.id);
      const past = isPast(d);
      const canDelete = entry && !past;
      h+=`<div class="btc" ${canDelete?`onclick="deleteBooking('${m.id}', '${dkey}')" style="cursor:pointer" title="Click to delete (admin)"`:''}>${statusPill(entry)}</div>`;
    });
    h+=`</div>`;
  });
  document.getElementById('rtbl').innerHTML=h;
}

// ── BOOK / CANCEL ─────────────────────────────────────────────────────────
async function bookSeat(sn){
  if(!S.user){toast('Please select your name first!','w');return;}
  const d=S.date; if(!d) return;
  const dkey=dk(d);
  const r=await apiBook(S.user,dkey,sn,'Office');
  if(r.success){bust(dkey);toast('✅ Seat '+sn+' booked!','s');await selDate(d);}
  else toast(r.message,'e');
}
async function setStatus(dateObjOrKey,status){
  if(!S.user){toast('Please select your name first!','w');return;}
  const dkey = typeof dateObjOrKey==='string' ? dateObjOrKey : dk(dateObjOrKey);
  const bkgs = await apiFetch(dkey);
  const existing = getEntry(bkgs, S.user);
  if(existing){ // remove old entry first (switching status or clearing)
    await apiCancel(S.user, dkey);
    bust(dkey);
  }
  if(!status){ // '' means "clear / no selection"
    toast('Cleared your plan for this day','i');
  } else if(status==='Office'){
    const fresh=await apiFetch(dkey);
    const taken=officeOf(fresh).map(b=>b.seat);
    let assigned=null;
    for(let s=1;s<=SEATS;s++){if(!taken.includes(s)){assigned=s;break;}}
    if(!assigned){toast('No seats available for this day!','e');await renderPanel();return;}
    const r=await apiBook(S.user,dkey,assigned,'Office');
    if(r.success){bust(dkey);toast('✅ Seat '+assigned+' booked!','s');}
    else toast(r.message,'e');
  } else {
    const r=await apiBook(S.user,dkey,null,status);
    if(r.success){bust(dkey);toast('✅ Marked as '+status,'s');}
    else toast(r.message,'e');
  }
  if(S.date && dk(S.date)===dkey) await selDate(S.date);
  else await renderPanel();
}
async function cancelSeat(sn, date=null){
  const d = date ? (typeof date === 'string' ? new Date(date) : date) : S.date;
  if(!d) return;
  if(isPast(d)){toast('Cannot cancel past bookings','w');return;}
  const dkey=dk(d);
  const r=await apiCancel(S.user,dkey);
  if(r.success){
    bust(dkey);
    toast('❌ Booking cancelled','i');
    if(S.date && dk(S.date)===dkey) await selDate(S.date);
    else await renderPanel();
  } else toast(r.message,'e');
}
async function deleteBooking(name, date, seat){
  if(!confirm(`Delete entry for ${name} on ${date}?`)) return;
  const r=await apiDelete(name, date, seat);
  if(r.success){
    bust(date);
    toast('🗑️ Entry deleted','i');
    if(S.date && dk(S.date)===date) await selDate(S.date);
    else await renderPanel();
  } else toast(r.message,'e');
}
async function quickBook(){
  const d=S.date; if(!d||!S.user) return;
  const dkey=dk(d);
  const bkgs=await apiFetch(dkey);
  if(getEntry(bkgs,S.user)){toast('You already have an entry on this day!','w');return;}
  const taken=officeOf(bkgs).map(b=>b.seat);
  for(let s=1;s<=SEATS;s++){if(!taken.includes(s)){await bookSeat(s);return;}}
  toast('No seats available!','e');
}
async function openWeekModal(){
  if(!S.user){toast('Please select your name first!','w');return;}
  const d=S.date||new Date();
  const wd=weekOf(d);
  const m=gm(S.user);
  showModal('📅 Book Entire Week',
    `<strong style="color:var(--sky)">${m?m.name:S.user}</strong> — booking a seat for all 5 working days of this week.<br><br>Already booked or full days will be skipped.`,
    [{l:'Book All 5 Days',c:'bp',fn:async()=>{closeModal();await bookWeek(d);}},{l:'Cancel',c:'bs',fn:closeModal}]
  );
}
async function bookWeek(date){
  const wd=weekOf(date); let n=0;
  for(const d of wd){
    if(isPast(d)) continue;
    const dkey=dk(d);
    const bkgs=await apiFetch(dkey);
    if(getEntry(bkgs,S.user)) continue;
    const taken=officeOf(bkgs).map(b=>b.seat);
    for(let s=1;s<=SEATS;s++){
      if(!taken.includes(s)){const r=await apiBook(S.user,dkey,s,'Office');if(r.success){bust(dkey);n++;}break;}
    }
  }
  toast('🎉 Booked '+n+' day(s) for the week!','s');
  if(S.date) await selDate(S.date);
}

// ── STATS ─────────────────────────────────────────────────────────────────
async function updateStats(){
  const d=S.date||new Date();
  const b=await apiFetch(dk(d));
  const off=officeOf(b).length;
  const wfh=b.filter(x=>x.status==='WFH').length;
  document.getElementById('stF').textContent=SEATS-off;
  document.getElementById('stB').textContent=off;
  document.getElementById('stW').textContent=wfh;
}

// ── REMINDER ──────────────────────────────────────────────────────────────
function checkReminder(){
  const now=new Date();
  const mins=now.getHours()*60+now.getMinutes();
  document.getElementById('rdesc').textContent=mins>=(16*60+30)?'Reminder sent at 4:30 PM today!':'Sends today at 4:30 PM IST';
}

// ── TABS ──────────────────────────────────────────────────────────────────
async function switchTab(tab,btn){
  S.tab=tab;
  document.querySelectorAll('.vp').forEach(p=>p.classList.remove('act'));
  document.querySelectorAll('.tab').forEach(t=>t.classList.remove('act'));
  document.getElementById('p'+tab[0]).classList.add('act');
  btn.classList.add('act');
  document.getElementById('wstrip').style.display = (tab==='week') ? 'none' : '';
  await renderPanel();
}
async function renderPanel(){
  if(S.tab==='floor') await renderFloor();
  else if(S.tab==='week') await renderWeek();
  else await renderRoster();
}

// ── MODAL ─────────────────────────────────────────────────────────────────
function showModal(title,body,actions){
  document.getElementById('mtitle').textContent=title;
  document.getElementById('mbody').innerHTML=body;
  document.getElementById('mact').innerHTML=actions.map((a,i)=>`<button class="btn ${a.c}" onclick="window._ma[${i}]()">${a.l}</button>`).join('');
  window._ma=actions.map(a=>a.fn);
  document.getElementById('mo').classList.add('act');
}
function closeModal(){document.getElementById('mo').classList.remove('act');}
document.getElementById('mo').addEventListener('click',function(e){if(e.target===this)closeModal();});

// ── TOAST ─────────────────────────────────────────────────────────────────
function toast(msg,type='i'){
  const icons={s:'✅',w:'⚠️',e:'❌',i:'ℹ️'};
  const t=document.createElement('div');
  t.className='toast '+type;
  t.innerHTML=`<span>${icons[type]}</span><span>${msg}</span>`;
  document.getElementById('tc').appendChild(t);
  setTimeout(()=>t.remove(),4000);
}

// ── USER SELECT ───────────────────────────────────────────────────────────
document.getElementById('usel').addEventListener('change',async function(){
  S.user=this.value;
  localStorage.setItem('ss_user',this.value);
  const m=gm(this.value);
  const av=document.getElementById('uav');
  if(m){av.textContent=m.ini;av.style.background=`linear-gradient(135deg,${m.color},${m.color}88)`;}
  else{av.textContent='?';av.style.background='linear-gradient(135deg,#0073AB,#00B5B1)';}
  updateDH();
  await renderPanel();
});

// ── CALENDAR NAV ──────────────────────────────────────────────────────────
document.getElementById('prevM').onclick=()=>{S.mo--;if(S.mo<0){S.mo=11;S.yr--;}renderCal();};
document.getElementById('nextM').onclick=()=>{S.mo++;if(S.mo>11){S.mo=0;S.yr++;}renderCal();};

// ── INIT ──────────────────────────────────────────────────────────────────
async function init(){
  // Restore saved user
  if(S.user){
    document.getElementById('usel').value=S.user;
    const m=gm(S.user);
    if(m){const av=document.getElementById('uav');av.textContent=m.ini;av.style.background=`linear-gradient(135deg,${m.color},${m.color}88)`;}
  }

  // Pick today or next Monday if weekend
  const today=new Date();
  let start=today;
  if(isWE(today)){
    start=new Date(today);
    start.setDate(today.getDate()+(today.getDay()===6?2:1));
  }
  S.yr=start.getFullYear();
  S.mo=start.getMonth();

  // Render calendar first so dates are visible immediately
  renderCal();

  // Then load today's data
  await selDate(start);
  checkReminder();
}

init();
</script>
</body>
</html>
"""

@app.route("/")
def home():
    return HTML_PAGE

@app.route("/api/bookings/<date_str>")
def api_get_bookings(date_str):
    rows = get_bookings_for_date(date_str)
    return jsonify({"date": date_str, "bookings": [{"name": r[0], "seat": r[1], "status": r[2]} for r in rows]})

@app.route("/api/book", methods=["POST"])
def api_book():
    data = request.json
    name, date_str = data.get("name"), data.get("date")
    status = data.get("status", "Office")  # Office / WFH / Travel / Leave
    seat = data.get("seat") if status == "Office" else None
    if status not in ("Office", "WFH", "Travel", "Leave"):
        return jsonify({"success": False, "message": "Invalid status"}), 400
    if not name or not date_str:
        return jsonify({"success": False, "message": "Missing fields"}), 400
    if status == "Office" and not seat:
        return jsonify({"success": False, "message": "Seat is required for Office bookings"}), 400
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM bookings WHERE name=%s AND date=%s", (name, date_str))
    if c.fetchone():
        conn.close()
        return jsonify({"success": False, "message": "You already have an entry on this day"}), 409
    if status == "Office":
        c.execute("SELECT * FROM bookings WHERE date=%s AND seat=%s AND status='Office'", (date_str, seat))
        if c.fetchone():
            conn.close()
            return jsonify({"success": False, "message": "That seat is already taken"}), 409
        c.execute("SELECT COUNT(*) FROM bookings WHERE date=%s AND status='Office'", (date_str,))
        if c.fetchone()[0] >= TOTAL_SEATS:
            conn.close()
            return jsonify({"success": False, "message": "No seats available"}), 409
    c.execute("INSERT INTO bookings (name, date, seat, status) VALUES (%s,%s,%s,%s)", (name, date_str, seat, status))
    conn.commit()
    conn.close()
    msg = f"Seat {seat} booked for {name}" if status == "Office" else f"{name} marked as {status}"
    return jsonify({"success": True, "message": msg})

@app.route("/api/cancel", methods=["POST"])
def api_cancel():
    data = request.json
    name, date_str = data.get("name"), data.get("date")
    conn = get_db()
    c = conn.cursor()
    c.execute("DELETE FROM bookings WHERE name=%s AND date=%s", (name, date_str))
    conn.commit()
    conn.close()
    return jsonify({"success": True, "message": "Booking cancelled"})

@app.route("/api/delete", methods=["POST"])
def api_delete():
    data = request.json
    name, date_str = data.get("name"), data.get("date")
    if not name or not date_str:
        return jsonify({"success": False, "message": "Missing fields"}), 400
    conn = get_db()
    c = conn.cursor()
    c.execute("DELETE FROM bookings WHERE name=%s AND date=%s", (name, date_str))
    conn.commit()
    conn.close()
    return jsonify({"success": True, "message": "Booking deleted"})

@app.route("/api/send-daily-reminder", methods=["POST"])
def manual_daily_reminder():
    """Lets you manually trigger the 9PM reminder early, for testing."""
    try:
        send_daily_seat_reminder()
        return jsonify({"success": True, "message": "Daily reminder sent!"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route("/api/send-monthly-roster", methods=["POST"])
def manual_monthly_roster():
    """Lets you manually trigger the monthly roster email, for testing."""
    try:
        send_monthly_roster()
        return jsonify({"success": True, "message": "Monthly roster sent!"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
